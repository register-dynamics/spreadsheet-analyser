from openpyxl import load_workbook

def check_cells_xlsx(file_path):
    wb = load_workbook(file_path)  
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        empty_cell_count = 0
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    print('------------------------\n')
                    print(f'cell value is {cell.value}')
                    print(f'cell fill is {cell.fill}')
                    print(f'cell fill patterntype: {cell.fill.patternType}')
                    print(f'indexed color: {cell.fill.fgColor.indexed}')
                    print('-------------------------\n')
                elif cell.value is None and empty_cell_count <= 1:
                    print('--------------------\n')
                    print(f'EMPTY CELL\n')
                    print(f'cell fill is {cell.fill}')
                    print(f'cell fill patterntype: {cell.fill.patternType}')
                    print(f'indexed color: {cell.fill.fgColor.indexed}')
                    print('---------------------\n')
                    empty_cell_count += 1
                    

def check_colored_cells_xlsx(file_path):
    wb = load_workbook(file_path)  
    has_color = False  
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                # Check if the cell has a fill and the fill is not "none"
                if cell.fill and cell.fill.patternType != 'none':
                    # Get the start color
                    color = cell.fill.start_color
                    # Check if the color is not default (e.g., transparent or white)
                    if color.index not in ['00000000', 'FFFFFFFF']:
                        print(f"Color {color.index} found in cell {cell.coordinate} of sheet '{sheet_name}'")
                        has_color = True

    return has_color

def check_styles_xlsx(file_path):
    workbook = load_workbook(file_path)
    has_styling = False

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.font.bold or cell.fill.patternType != 'none' or cell.fill.start_color.index != '00000000':
                    print(f"Styling found in cell {cell.coordinate} of sheet '{sheet_name}'")
                    has_styling = True

    return has_styling
    
if __name__ == '__main__':
    # Example usage:
    file_path = "/Users/gibbons/registerdynamics/python_projects/url_tester/test_sheet_colored.xlsx" 
    '''merged_cells = check_merged_cells_xlsx(file_path)
    if merged_cells:
        print("Merged cells detected:")
        for sheet_name, merged_range in merged_cells:
            print(f"- {merged_range} in sheet '{sheet_name}'")
    else:
        print("No merged cells found.")'''
    check_cells_xlsx(file_path)